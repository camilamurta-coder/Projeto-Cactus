#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera os dados do painel "Jornada rumo a medalha" a partir de:
  - planilha de gestao (abas Painel / Checklist / Acesso)
  - export de gamificacao (CSV por aluno)

Saidas:
  - data.json                          -> dados AGREGADOS (publico, vai pro site)
  - interno/gamificacao_detalhado.csv  -> lista por aluno, com nome (uso interno,
                                           NUNCA commitar / publicar - ver .gitignore)

Uso:
    python build_data.py [caminho_xlsx] [caminho_csv_gamificacao]

Se os caminhos nao forem passados, usa os arquivos mais recentes enviados
(constantes abaixo) como padrao.
"""

from __future__ import annotations

import csv
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

XLSX_PADRAO = r"C:\Users\Camila\Downloads\2026_Acompanhamento_JornadaOBMEP (1).xlsx"
CSV_PADRAO = r"C:\Users\Camila\Downloads\gamificacao (3).csv"

AQUI = Path(__file__).parent
INTERNO = AQUI / "interno"


def norm(s) -> str:
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().upper()


def num(v, default=0):
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def data_iso(v):
    if v is None:
        return None
    try:
        return v.date().isoformat()
    except AttributeError:
        return str(v)


# --------------------------------------------------------------------------- #
# PLANILHA DE GESTAO
# --------------------------------------------------------------------------- #

def ler_painel(ws) -> dict:
    linhas = list(ws.iter_rows(values_only=True))

    def achar(rotulo):
        for row in linhas:
            if row and row[0] and norm(row[0]) == norm(rotulo):
                return row
        return None

    meta = {
        "hoje": data_iso(achar("Hoje")[1]) if achar("Hoje") else None,
        "divulgacao_locais": data_iso(achar("Divulgação dos locais de prova")[1])
        if achar("Divulgação dos locais de prova") else None,
        "prova_data": data_iso(achar("Prova da 2ª fase")[1]) if achar("Prova da 2ª fase") else None,
        "dias_corridos": achar("Dias corridos até a prova")[1] if achar("Dias corridos até a prova") else None,
    }

    # blocos do checklist (tabela "Bloco | Entregáveis | Concluídos | % concluído | Atrasados")
    blocos = []
    capturando = False
    for row in linhas:
        if row[0] == "Bloco" and row[1] == "Entregáveis":
            capturando = True
            continue
        if capturando:
            if row[0] is None:
                break
            blocos.append({
                "bloco": row[0],
                "entregaveis": int(num(row[1])),
                "concluidos": int(num(row[2])),
                "pct_concluido": num(row[3]),
                "atrasados": int(num(row[4])),
            })

    acessos_geral = {
        "meta_pct": num(achar("Meta de acessos (% dos estudantes na trilha)")[1]),
        "estudantes_n1n2": int(num(achar("Estudantes N1+N2 na trilha")[1])),
        "acesso_registrado": int(num(achar("Acesso registrado")[1])),
        "pct_alcancado": num(achar("% de acesso alcançado")[1]),
        "municipios_prioridade_alta": int(num(achar("Municípios com prioridade Alta")[1])),
        "municipios_sem_consultor": int(num(achar("Municípios sem consultor definido")[1])),
        "escolas_municipais": int(num(achar("Escolas municipais na trilha")[1])),
        "municipios_base_auditar": int(num(achar("Municípios com base ainda a auditar")[1])),
    }

    meta_geral = {
        "municipios_jornada": int(num(achar("Municípios na Jornada")[1])),
        "meta_municipios_medalhista": int(num(achar("META: municípios com ao menos 1 medalhista")[1])),
        "municipios_meta_atingida": int(num(achar("Municípios com a meta atingida")[1])),
        "medalhistas_meta_somada": int(num(achar("Medalhistas: meta somada")[1])),
        "medalhistas_f2_realizado": int(num(achar("Medalhistas Fase 2: realizado")[1])),
        "medalhistas_f1_base": int(num(achar("Medalhistas Fase 1 (linha de base)")[1])),
    }

    # tabela de consultores
    consultores = []
    capturando = False
    for row in linhas:
        if row[0] == "Consultor":
            capturando = True
            continue
        if capturando:
            if row[0] is None:
                break
            consultores.append({
                "consultor": row[0],
                "municipios": int(num(row[1])),
                "estudantes_n1n2": int(num(row[2])),
                "meta_acesso": num(row[3]),
                "acesso_registrado": int(num(row[4])),
                "atingimento_pct": num(row[5]),
                "municipios_com_medalhista": int(num(row[6])),
                "municipios_sem_medalhista_f1": int(num(row[7])),
            })

    return {
        "titulo": linhas[0][0] if linhas else None,
        "subtitulo": linhas[1][0] if len(linhas) > 1 else None,
        "datas_chave": meta,
        "checklist_progresso": blocos,
        "acessos_geral": acessos_geral,
        "meta_geral": meta_geral,
        "consultores": consultores,
    }


def ler_checklist(ws) -> list[dict]:
    linhas = list(ws.iter_rows(values_only=True))
    itens = []
    cabecalho_visto = False
    for row in linhas:
        if row[0] == "ID":
            cabecalho_visto = True
            continue
        if not cabecalho_visto:
            continue
        if row[0] is None:
            continue
        itens.append({
            "id": int(row[0]),
            "bloco": row[1],
            "entregavel": row[2],
            "pronto_quando": row[3],
            "responsavel": row[4],
            "prazo": data_iso(row[5]),
            "status": row[6],
            "data_conclusao": data_iso(row[7]),
            "observacoes": row[8],
            "atrasado": row[9] == "Sim",
        })
    return itens


def ler_acesso(ws) -> list[dict]:
    linhas = list(ws.iter_rows(values_only=True))
    municipios = []
    cabecalho_idx = None
    for i, row in enumerate(linhas):
        if row[0] == "Município":
            cabecalho_idx = i
            break
    if cabecalho_idx is None:
        return municipios

    for row in linhas[cabecalho_idx + 1:]:
        nome = row[0]
        if nome is None or nome in ("Total",):
            continue
        if norm(nome) in ("TRAVAS DE INTEGRIDADE",):
            break
        municipios.append({
            "municipio": nome,
            "uf": row[1],
            "consultor": row[2],
            "estudantes_n1n2": int(num(row[3])),
            "escolas_municipais": int(num(row[4])),
            "meta_pct": num(row[5]),
            "meta_acesso_abs": num(row[6]),
            "acesso_registrado": int(num(row[7])),
            "pct_acesso": num(row[8]),
            "medalhistas_f1": int(num(row[9])),
            "meta_medalhistas_f2": int(num(row[10])),
            "medalhistas_f2_realizado": int(num(row[11])) if row[11] is not None else 0,
            "meta_atingida": row[12],
            "prioridade": row[13],
            "observacoes": row[14],
            "farol": row[33] if len(row) > 33 else None,
            "motivo_farol": row[34] if len(row) > 34 else None,
        })
    return municipios


# --------------------------------------------------------------------------- #
# GAMIFICACAO (CSV por aluno)
# --------------------------------------------------------------------------- #

def ler_gamificacao(caminho: Path) -> list[dict]:
    alunos = []
    with open(caminho, encoding="utf-8-sig", errors="replace", newline="") as f:
        leitor = csv.DictReader(f, delimiter=";")
        for row in leitor:
            turma = (row.get("Turma") or "").strip()
            if " - " in turma:
                nivel_turma, municipio = turma.split(" - ", 1)
            else:
                nivel_turma, municipio = turma, ""
            alunos.append({
                "turma": turma,
                "nivel_turma": nivel_turma.strip(),
                "municipio": municipio.strip(),
                "municipio_norm": norm(municipio),
                "aluno": (row.get("Aluno") or "").strip(),
                "nivel": int(num(row.get("Nível"))),
                "pontos": int(num(row.get("Pontos"))),
                "conquistas": int(num(row.get("Conquistas"))),
            })

    # remove duplicatas exatas (mesmo aluno/turma/nivel/pontos/conquistas) -
    # artefato comum de export do LMS, nao estudantes distintos
    vistos = set()
    dedup = []
    for a in alunos:
        chave = (norm(a["aluno"]), a["turma"], a["nivel"], a["pontos"], a["conquistas"])
        if chave in vistos:
            continue
        vistos.add(chave)
        dedup.append(a)
    return dedup


def agregar_gamificacao(alunos: list[dict], municipios_jornada: list[str]) -> dict:
    def resumo(grupo):
        total = len(grupo)
        engajados = sum(1 for a in grupo if a["pontos"] > 0)
        pontos = [a["pontos"] for a in grupo]
        niveis = [a["nivel"] for a in grupo]
        return {
            "total_alunos": total,
            "alunos_engajados": engajados,
            "pct_engajados": round(engajados / total, 4) if total else 0,
            "pontos_totais": sum(pontos),
            "pontos_medio_engajados": round(sum(pontos) / engajados, 1) if engajados else 0,
            "nivel_medio": round(sum(niveis) / total, 2) if total else 0,
            "conquistas_totais": sum(a["conquistas"] for a in grupo),
            "maior_pontuacao": max(pontos) if pontos else 0,
        }

    por_municipio = {}
    for a in alunos:
        por_municipio.setdefault(a["municipio_norm"], {"nome_exibicao": a["municipio"], "linhas": []})
        por_municipio[a["municipio_norm"]]["linhas"].append(a)

    lista_municipios = []
    for chave, dados in por_municipio.items():
        r = resumo(dados["linhas"])
        r["municipio"] = dados["nome_exibicao"]
        por_nivel = {}
        for a in dados["linhas"]:
            por_nivel.setdefault(a["nivel_turma"], []).append(a)
        r["por_nivel_turma"] = {niv: resumo(g) for niv, g in por_nivel.items()}
        lista_municipios.append(r)

    # inclui, com zero, os municipios da Jornada que ainda nao tem nenhum
    # estudante na base de gamificacao (para aparecerem na lista tambem)
    vazio = resumo([])
    presentes = {norm(r["municipio"]) for r in lista_municipios}
    for nome in municipios_jornada:
        if norm(nome) not in presentes:
            r = dict(vazio)
            r["municipio"] = nome
            r["por_nivel_turma"] = {}
            r["sem_inscritos"] = True
            lista_municipios.append(r)

    # decrescente por engajamento; empates (inclusive 0/sem inscritos) ficam
    # em ordem alfabetica entre si
    lista_municipios.sort(key=lambda r: norm(r["municipio"]))
    lista_municipios.sort(key=lambda r: r["pct_engajados"], reverse=True)

    geral = resumo(alunos)
    geral["municipios_com_inscritos"] = sum(1 for r in lista_municipios if r["total_alunos"] > 0)
    geral["municipios_com_ativos"] = sum(1 for r in lista_municipios if r["alunos_engajados"] > 0)
    geral["total_municipios_jornada"] = len(municipios_jornada)

    return {
        "geral": geral,
        "por_municipio": lista_municipios,
    }


# --------------------------------------------------------------------------- #
# EXECUCAO
# --------------------------------------------------------------------------- #

def main() -> int:
    caminho_xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(XLSX_PADRAO)
    caminho_csv = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(CSV_PADRAO)

    if not caminho_xlsx.exists():
        print(f"! planilha nao encontrada: {caminho_xlsx}")
        return 1
    if not caminho_csv.exists():
        print(f"! csv nao encontrado: {caminho_csv}")
        return 1

    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    painel = ler_painel(wb["Painel"])
    checklist_itens = ler_checklist(wb["Checklist"])
    municipios_acesso = ler_acesso(wb["Acesso"])
    municipios_acesso.sort(key=lambda m: norm(m["municipio"]))

    alunos = ler_gamificacao(caminho_csv)
    gamificacao = agregar_gamificacao(
        alunos, municipios_jornada=[m["municipio"] for m in municipios_acesso]
    )

    saida = {
        "gerado_em_arquivo": {
            "planilha": caminho_xlsx.name,
            "csv_gamificacao": caminho_csv.name,
        },
        "painel": painel,
        "checklist_itens": checklist_itens,
        "municipios_acesso": municipios_acesso,
        "gamificacao": gamificacao,
    }

    (AQUI / "data.json").write_text(
        json.dumps(saida, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"data.json gerado com {len(municipios_acesso)} municipios e "
          f"{len(gamificacao['por_municipio'])} municipios com dados de gamificacao "
          f"({len(alunos)} alunos, sem nomes no arquivo publico).")

    # lista interna (COM nome do aluno) - nunca vai pro site/git
    INTERNO.mkdir(exist_ok=True)
    caminho_interno = INTERNO / "gamificacao_detalhado.csv"
    with open(caminho_interno, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Municipio", "Nivel_Turma", "Aluno", "Nivel", "Pontos", "Conquistas"])
        for a in sorted(alunos, key=lambda a: (a["municipio"], a["nivel_turma"], -a["pontos"])):
            w.writerow([a["municipio"], a["nivel_turma"], a["aluno"], a["nivel"],
                        a["pontos"], a["conquistas"]])
    print(f"lista interna (com nomes) gerada em {caminho_interno} - NAO sera publicada")

    # monta o index.html final embutindo os dados no template
    template = (AQUI / "index.template.html").read_text(encoding="utf-8")
    bloco_json = json.dumps(saida, ensure_ascii=False).replace("</", "<\\/")
    html_final = template.replace("__DASHBOARD_DATA__", bloco_json)
    (AQUI / "index.html").write_text(html_final, encoding="utf-8")
    print("index.html gerado a partir do template.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
